"""Excel .xlsx import for AI Products (0 AI credits)."""

from __future__ import annotations

import io
from typing import Any

from services.products.import_parser import normalize_import_row


def parse_xlsx_bytes(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [str(cell or "").strip() for cell in header_row]
    parsed_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(rows_iter, start=2):
        raw: dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = row[col_idx] if col_idx < len(row) else None
            raw[header] = value
        if not any(str(v or "").strip() for v in raw.values()):
            continue
        parsed_rows.append(normalize_import_row(raw, row_number=idx))
    return parsed_rows


def build_xlsx_template_bytes() -> bytes:
    from openpyxl import Workbook

    from services.products.import_parser import import_template_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = import_template_headers()
    ws.append(headers)
    ws.append(
        [
            "Rose Lipstick",
            "45",
            "AED",
            "S,M",
            "Rose,Nude",
            "Matte finish",
            "Lip color",
            "in_stock",
            "https://example.com/img1.jpg",
            "",
            "",
            "https://tiktok.com/@shop",
            "https://instagram.com/shop",
            "",
            "https://shop.example.com/item",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
